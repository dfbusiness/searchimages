#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SearchImages
-------------
Ferramenta para gerar planilhas comerciais com imagens reais de produtos a partir
 de links de fornecedores constantes em uma planilha padrão.

Fluxo principal:
1. Lê uma planilha de memória de cálculo.
2. Opcionalmente lê uma tabela de descrição/unidade/quantidade em PDF.
3. Acessa os links de referência dos produtos.
4. Para Mercado Livre, tenta primeiro a API pública.
5. Para demais sites ou falhas, usa Playwright com navegador real.
6. Baixa/captura a imagem principal real do produto.
7. Gera planilha final com imagens, marcas, valores e log de conferência.

Importante:
- A ferramenta não usa imagens genéricas, ícones, placeholders ou imagens criadas.
- Quando não consegue capturar a imagem real com segurança, deixa em branco e marca REVISAR.
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import re
import shutil
import sys
import time
from dataclasses import dataclass, asdict
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlparse

import openpyxl
import requests
import yaml
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from PIL import Image, ImageOps

try:
    import pdfplumber
except Exception:  # pragma: no cover
    pdfplumber = None

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
except Exception:  # pragma: no cover
    sync_playwright = None
    PlaywrightTimeoutError = Exception


USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0 Safari/537.36"
)

HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
}

MLB_ID_RE = re.compile(r"\b(MLB\d{6,})\b", re.IGNORECASE)


@dataclass
class ProductRow:
    item: int
    description: str
    unit: str
    quantity: float
    unit_price: float
    source_url: str
    memory_description: str = ""


@dataclass
class CaptureResult:
    item: int
    status: str
    source: str
    brand: str
    image_path: str
    title: str
    observations: str
    url: str


# ---------------------------------------------------------------------------
# Utilidades gerais
# ---------------------------------------------------------------------------


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%H:%M:%S",
    )


def load_config(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def to_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def safe_filename(name: str, fallback: str = "image") -> str:
    name = re.sub(r"[^A-Za-z0-9_.-]+", "_", name).strip("_")
    return name or fallback


def cm_to_pixels(cm: float, dpi: int = 96) -> int:
    return int(cm / 2.54 * dpi)


def image_to_png_bytes(img: Image.Image) -> bytes:
    out = BytesIO()
    img.save(out, format="PNG")
    return out.getvalue()


def normalize_image_file(input_path: Path, output_path: Path, max_height_px: int = 320) -> None:
    img = Image.open(input_path)
    img = ImageOps.exif_transpose(img).convert("RGB")
    if img.height > max_height_px:
        ratio = max_height_px / img.height
        new_size = (max(1, int(img.width * ratio)), max_height_px)
        img = img.resize(new_size, Image.LANCZOS)
    img.save(output_path, "PNG", optimize=True)


def download_image(url: str, output_path: Path, timeout: int = 30) -> bool:
    if not url or not url.startswith(("http://", "https://")):
        return False
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout)
        r.raise_for_status()
        content_type = r.headers.get("content-type", "").lower()
        if "image" not in content_type and len(r.content) < 5000:
            return False
        img = Image.open(BytesIO(r.content))
        img = ImageOps.exif_transpose(img).convert("RGB")
        # Evita salvar imagens muito pequenas, ícones ou logos.
        if img.width < 160 or img.height < 160:
            return False
        img.save(output_path, "PNG", optimize=True)
        return True
    except Exception as exc:
        logging.debug("Falha ao baixar imagem %s: %s", url, exc)
        return False


# ---------------------------------------------------------------------------
# Leitura da memória de cálculo e PDF
# ---------------------------------------------------------------------------


def cell(ws, row: int, col_letter: str) -> Any:
    return ws.cell(row=row, column=column_index_from_string(col_letter)).value


def read_memory_excel(config: Dict[str, Any]) -> List[ProductRow]:
    excel_path = Path(config["input"]["excel_path"])
    sheet_name = config["input"].get("sheet_name")
    first_row = int(config["input"].get("first_data_row", 2))
    last_row = int(config["input"].get("last_data_row", first_row))
    cols = config["columns"]

    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows: List[ProductRow] = []
    for row in range(first_row, last_row + 1):
        desc_memory = clean_text(cell(ws, row, cols.get("description_memory", "B")))
        url = clean_text(cell(ws, row, cols.get("url", "U")))
        if not desc_memory and not url:
            continue

        raw_item = clean_text(cell(ws, row, cols.get("item", "A")))
        m = re.search(r"(\d+)", raw_item)
        item = int(m.group(1)) if m else len(rows) + 1

        quantity = to_float(cell(ws, row, cols.get("quantity", "F")))
        unit_price = to_float(cell(ws, row, cols.get("unit_price", "S")))

        rows.append(
            ProductRow(
                item=item,
                description=desc_memory,
                unit="",
                quantity=quantity,
                unit_price=unit_price,
                source_url=url,
                memory_description=desc_memory,
            )
        )
    return rows


def extract_pdf_table(pdf_path: Path) -> Dict[int, Tuple[str, str, float]]:
    """Extrai tabela Item / Descrição / Unidade / Qtde do PDF quando possível.

    Retorna dict: item -> (descricao, unidade, qtde)
    """
    results: Dict[int, Tuple[str, str, float]] = {}
    if not pdf_path.exists() or pdfplumber is None:
        return results

    all_text = ""
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            all_text += "\n" + text

    # Captura linhas da tabela do TR. Algumas descrições quebram linha; por isso usamos
    # uma estratégia por item, olhando do número atual até o próximo número.
    start = all_text.find("ITEM DESCRIÇÃO")
    end = all_text.find("4 - GARANTIA")
    table_text = all_text[start:end] if start != -1 and end != -1 else all_text
    lines = [clean_text(x) for x in table_text.splitlines() if clean_text(x)]

    # Une possíveis quebras de linha até encontrar unidade/qtde no final.
    buffer = ""
    for ln in lines:
        if re.match(r"^\d+\s+", ln):
            if buffer:
                _parse_pdf_table_line(buffer, results)
            buffer = ln
        else:
            # ignora cabeçalhos, mas concatena continuação de descrição
            if buffer and not ln.lower().startswith(("item ", "tabela", "unidade", "qtde")):
                buffer += " " + ln
    if buffer:
        _parse_pdf_table_line(buffer, results)

    return results


def _parse_pdf_table_line(line: str, results: Dict[int, Tuple[str, str, float]]) -> None:
    # Exemplos finais: "Conjunto 4", "Peça 1", "peça 100", "litro 40", "Pacote 30"
    m = re.match(r"^(\d+)\s+(.+?)\s+(Conjunto|Peça|peça|Pacote|litro)\s+(\d+(?:[.,]\d+)?)$", line, re.IGNORECASE)
    if not m:
        return
    item = int(m.group(1))
    desc = clean_text(m.group(2))
    unit = clean_text(m.group(3))
    qty = to_float(m.group(4))
    results[item] = (desc, unit, qty)


def merge_pdf_data(rows: List[ProductRow], config: Dict[str, Any]) -> List[ProductRow]:
    pdf_cfg = config.get("pdf_table", {})
    if not pdf_cfg.get("enabled", False):
        return rows
    pdf_path = Path(config["input"].get("pdf_path", ""))
    pdf_data = extract_pdf_table(pdf_path)
    if not pdf_data:
        logging.warning("Não foi possível extrair tabela do PDF. Usando descrições da memória de cálculo.")
        return rows

    for row in rows:
        if row.item in pdf_data:
            desc, unit, qty = pdf_data[row.item]
            row.description = desc
            row.unit = unit
            if qty:
                row.quantity = qty
    return rows


# ---------------------------------------------------------------------------
# Identificação de marca
# ---------------------------------------------------------------------------


def brand_from_title_or_url(title: str, url: str = "") -> str:
    text = f"{title} {url}".lower()
    known = {
        "tramontina": "Tramontina",
        "porteña": "Porteña",
        "portena": "Porteña",
        "mk": "MK",
        "oxford": "Oxford",
        "bormioli": "Bormioli Rocco",
        "deca": "Deca",
        "snob": "Snob",
        "wd-40": "WD-40",
        "wd40": "WD-40",
        "azulim": "Azulim",
        "união": "União",
        "uniao": "União",
        "3m": "3M",
        "lixex": "Lixex",
        "premisse": "Premisse",
        "condor": "Condor",
        "lar plasticos": "Lar Plásticos",
        "lar plásticos": "Lar Plásticos",
        "max ferramentas": "Max Ferramentas",
        "sanches": "Sanches",
        "mecflux": "Mecflux",
    }
    for key, value in known.items():
        if key in text:
            return value
    return "Não identificada no link"


def brand_from_attributes(attrs: Any, title: str, url: str) -> str:
    if isinstance(attrs, list):
        for attr in attrs:
            name = clean_text(attr.get("name", "")).lower()
            attr_id = clean_text(attr.get("id", "")).upper()
            value = clean_text(attr.get("value_name", ""))
            if value and ("marca" in name or attr_id == "BRAND"):
                return value
    return brand_from_title_or_url(title, url)


# ---------------------------------------------------------------------------
# Mercado Livre API
# ---------------------------------------------------------------------------


def extract_mlb_ids(url: str) -> List[str]:
    return list(dict.fromkeys(x.upper() for x in MLB_ID_RE.findall(url or "")))


def mercadolivre_api_capture(row: ProductRow, images_dir: Path) -> Optional[CaptureResult]:
    ids = extract_mlb_ids(row.source_url)
    if not ids:
        return None

    # Prioriza IDs de anúncio com mais dígitos quando existirem dois IDs no link.
    ids = sorted(ids, key=len, reverse=True)
    item_id = ids[0]
    api_url = f"https://api.mercadolibre.com/items/{item_id}"
    try:
        r = requests.get(api_url, headers=HEADERS, timeout=30)
        if r.status_code == 404 and len(ids) > 1:
            item_id = ids[1]
            api_url = f"https://api.mercadolibre.com/items/{item_id}"
            r = requests.get(api_url, headers=HEADERS, timeout=30)
        r.raise_for_status()
        data = r.json()
    except Exception as exc:
        return CaptureResult(row.item, "REVISAR", "mercadolivre_api", "Não identificada no link", "", "", f"Falha API ML: {exc}", row.source_url)

    title = clean_text(data.get("title", ""))
    brand = brand_from_attributes(data.get("attributes"), title, row.source_url)
    pictures = data.get("pictures") or []
    image_url = ""
    if pictures:
        image_url = pictures[0].get("secure_url") or pictures[0].get("url") or ""
    elif data.get("thumbnail"):
        image_url = data.get("thumbnail")

    out = images_dir / f"item_{row.item:02d}.png"
    if image_url and download_image(image_url, out):
        return CaptureResult(row.item, "OK", "mercadolivre_api", brand, str(out), title, "Imagem obtida pela API Mercado Livre", row.source_url)

    return CaptureResult(row.item, "REVISAR", "mercadolivre_api", brand, "", title, "API retornou dados, mas a imagem não pôde ser baixada", row.source_url)


# ---------------------------------------------------------------------------
# Playwright visual
# ---------------------------------------------------------------------------


def _candidate_images_from_page(page) -> List[Tuple[str, int, int, str]]:
    """Retorna lista de imagens candidatas: (src, width, height, alt)."""
    js = r"""
    () => {
      const out = [];
      const imgs = Array.from(document.images || []);
      for (const img of imgs) {
        const rect = img.getBoundingClientRect();
        const src = img.currentSrc || img.src || img.getAttribute('data-src') || '';
        const alt = img.alt || '';
        const width = Math.round(rect.width || img.naturalWidth || 0);
        const height = Math.round(rect.height || img.naturalHeight || 0);
        if (!src) continue;
        out.push({src, width, height, alt, top: Math.round(rect.top), left: Math.round(rect.left)});
      }
      return out;
    }
    """
    data = page.evaluate(js)
    candidates = []
    for item in data:
        src = item.get("src", "")
        width = int(item.get("width") or 0)
        height = int(item.get("height") or 0)
        alt = item.get("alt", "") or ""
        top = int(item.get("top") or 999999)
        # Regras conservadoras contra logos/ícones.
        if width < 180 or height < 180:
            continue
        if any(bad in src.lower() for bad in ["logo", "sprite", "icon", "avatar", "payment", "banner"]):
            continue
        score = width * height - max(top, 0) * 50
        candidates.append((src, width, height, alt, score))
    candidates.sort(key=lambda x: x[4], reverse=True)
    return [(a, b, c, d) for a, b, c, d, _ in candidates]


def _brand_from_page_meta(page, title: str, url: str) -> str:
    # Tenta structured data JSON-LD primeiro.
    try:
        scripts = page.locator('script[type="application/ld+json"]').all_inner_texts()
        for raw in scripts:
            try:
                data = json.loads(raw)
            except Exception:
                continue
            nodes = data if isinstance(data, list) else [data]
            for node in nodes:
                if not isinstance(node, dict):
                    continue
                brand = node.get("brand")
                if isinstance(brand, dict):
                    val = clean_text(brand.get("name"))
                    if val:
                        return val
                elif isinstance(brand, str) and brand.strip():
                    return clean_text(brand)
    except Exception:
        pass
    return brand_from_title_or_url(title, url)


def playwright_capture(row: ProductRow, images_dir: Path, config: Dict[str, Any], browser_context=None) -> CaptureResult:
    if sync_playwright is None and browser_context is None:
        return CaptureResult(row.item, "REVISAR", "playwright", "Não identificada no link", "", "", "Playwright não instalado", row.source_url)

    timeout_ms = int(config.get("behavior", {}).get("timeout_ms", 45000))
    headless = bool(config.get("behavior", {}).get("headless", False))

    own_context = False
    pw = None
    browser = None
    context = browser_context
    try:
        if context is None:
            own_context = True
            pw = sync_playwright().start()
            browser = pw.chromium.launch(headless=headless)
            context = browser.new_context(user_agent=USER_AGENT, viewport={"width": 1366, "height": 900}, locale="pt-BR")

        page = context.new_page()
        page.goto(row.source_url, wait_until="domcontentloaded", timeout=timeout_ms)
        try:
            page.wait_for_load_state("networkidle", timeout=12000)
        except Exception:
            pass
        time.sleep(2)
        title = clean_text(page.title())
        brand = _brand_from_page_meta(page, title, row.source_url)
        candidates = _candidate_images_from_page(page)
        out = images_dir / f"item_{row.item:02d}.png"

        for src, width, height, alt in candidates[:8]:
            if download_image(src, out):
                page.close()
                return CaptureResult(row.item, "OK", "playwright_image", brand, str(out), title, f"Imagem candidata baixada: {width}x{height}; alt={alt[:80]}", row.source_url)

        # Fallback conservador: screenshot da maior imagem candidata.
        if config.get("behavior", {}).get("screenshot_fallback", True):
            locator = page.locator("img").filter(has_not_text="")
            count = locator.count()
            best_idx = None
            best_area = 0
            for i in range(min(count, 80)):
                try:
                    box = locator.nth(i).bounding_box(timeout=2000)
                except Exception:
                    box = None
                if not box:
                    continue
                w, h = int(box.get("width", 0)), int(box.get("height", 0))
                if w >= 180 and h >= 180 and w * h > best_area:
                    best_area = w * h
                    best_idx = i
            if best_idx is not None:
                locator.nth(best_idx).screenshot(path=str(out), timeout=8000)
                page.close()
                return CaptureResult(row.item, "OK", "playwright_screenshot", brand, str(out), title, "Screenshot da maior imagem principal visível", row.source_url)

        page.close()
        return CaptureResult(row.item, "REVISAR", "playwright", brand, "", title, "Nenhuma imagem principal capturada com segurança", row.source_url)
    except Exception as exc:
        return CaptureResult(row.item, "REVISAR", "playwright", "Não identificada no link", "", "", f"Erro Playwright: {exc}", row.source_url)
    finally:
        if own_context:
            try:
                context.close()
            except Exception:
                pass
            try:
                browser.close()
            except Exception:
                pass
            try:
                pw.stop()
            except Exception:
                pass


def capture_all_images(rows: List[ProductRow], config: Dict[str, Any]) -> List[CaptureResult]:
    images_dir = Path(config["output"]["images_dir"])
    ensure_dir(images_dir)
    results: List[CaptureResult] = []

    prefer_ml = bool(config.get("behavior", {}).get("prefer_mercadolivre_api", True))
    use_pw = bool(config.get("behavior", {}).get("use_playwright_fallback", True))

    # Usa um único navegador para acelerar fallback visual.
    pw = browser = context = None
    if use_pw and sync_playwright is not None:
        try:
            headless = bool(config.get("behavior", {}).get("headless", False))
            pw = sync_playwright().start()
            browser = pw.chromium.launch(headless=headless)
            context = browser.new_context(user_agent=USER_AGENT, viewport={"width": 1366, "height": 900}, locale="pt-BR")
        except Exception as exc:
            logging.warning("Não foi possível iniciar Playwright: %s", exc)
            context = None

    try:
        for row in rows:
            logging.info("Item %02d: coletando imagem", row.item)
            result: Optional[CaptureResult] = None

            if prefer_ml and "mercadolivre" in row.source_url.lower():
                result = mercadolivre_api_capture(row, images_dir)
                if result and result.status == "OK":
                    results.append(result)
                    continue
                # Mantém observação da API, mas tenta visual.
                previous_obs = result.observations if result else ""
            else:
                previous_obs = ""

            if use_pw:
                pw_result = playwright_capture(row, images_dir, config, browser_context=context)
                if previous_obs and pw_result.status != "OK":
                    pw_result.observations = previous_obs + " | " + pw_result.observations
                result = pw_result
            elif result is None:
                result = CaptureResult(row.item, "REVISAR", "none", "Não identificada no link", "", "", "Coleta visual desativada", row.source_url)

            results.append(result)
    finally:
        if context:
            context.close()
        if browser:
            browser.close()
        if pw:
            pw.stop()

    return results


# ---------------------------------------------------------------------------
# Geração da planilha final
# ---------------------------------------------------------------------------


def currency_format(cell) -> None:
    cell.number_format = 'R$ #,##0.00'


def build_output_workbook(rows: List[ProductRow], captures: List[CaptureResult], config: Dict[str, Any]) -> Path:
    output_path = Path(config["output"]["workbook_path"])
    ensure_dir(output_path.parent)

    cap_by_item = {c.item: c for c in captures}
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha"

    headers = ["Item", "Descrição do Item", "Imagem", "UN", "Qtde", "Valor Unitário", "Valor Total"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, len(headers) + 1):
        cell_obj = ws.cell(1, col)
        cell_obj.fill = header_fill
        cell_obj.font = header_font
        cell_obj.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_obj.border = border

    max_height_cm = float(config["output"].get("image_max_height_cm", 3.0))
    max_height_px = cm_to_pixels(max_height_cm)

    for idx, row in enumerate(rows, start=2):
        cap = cap_by_item.get(row.item)
        brand = cap.brand if cap else "Não identificada no link"
        desc = f"{row.description}\nMarca: {brand}"
        ws.cell(idx, 1).value = row.item
        ws.cell(idx, 2).value = desc
        ws.cell(idx, 4).value = row.unit or ""
        ws.cell(idx, 5).value = row.quantity
        ws.cell(idx, 6).value = row.unit_price
        ws.cell(idx, 7).value = f"=E{idx}*F{idx}"

        for col in range(1, 8):
            c = ws.cell(idx, col)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = border
        currency_format(ws.cell(idx, 6))
        currency_format(ws.cell(idx, 7))

        # Altura aproximada para 3 cm; Excel usa pontos. 1 cm ≈ 28,35 pt.
        ws.row_dimensions[idx].height = max_height_cm * 28.35 + 8

        if cap and cap.status == "OK" and cap.image_path and Path(cap.image_path).exists():
            img_path = Path(cap.image_path)
            try:
                # Normaliza proporção/tamanho antes de inserir.
                normalized = img_path.with_name(img_path.stem + "_excel.png")
                normalize_image_file(img_path, normalized, max_height_px=max_height_px)
                xl_img = XLImage(str(normalized))
                if xl_img.height > max_height_px:
                    ratio = max_height_px / xl_img.height
                    xl_img.height = max_height_px
                    xl_img.width = int(xl_img.width * ratio)
                ws.add_image(xl_img, f"C{idx}")
            except Exception as exc:
                logging.warning("Falha ao inserir imagem do item %s: %s", row.item, exc)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"
    widths = {"A": 9, "B": 58, "C": 22, "D": 12, "E": 10, "F": 16, "G": 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Aba de fontes/log
    if config["output"].get("create_sources_sheet", True):
        src = wb.create_sheet("Fontes")
        src_headers = ["Item", "Link", "Marca identificada", "Arquivo da imagem", "Status", "Fonte de captura", "Título", "Observações"]
        src.append(src_headers)
        for col in range(1, len(src_headers) + 1):
            c = src.cell(1, col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        for idx, row in enumerate(rows, start=2):
            cap = cap_by_item.get(row.item)
            src.cell(idx, 1).value = row.item
            src.cell(idx, 2).value = row.source_url
            if config["output"].get("add_hyperlinks", True) and row.source_url:
                src.cell(idx, 2).hyperlink = row.source_url
                src.cell(idx, 2).style = "Hyperlink"
            src.cell(idx, 3).value = cap.brand if cap else ""
            src.cell(idx, 4).value = cap.image_path if cap else ""
            src.cell(idx, 5).value = cap.status if cap else "REVISAR"
            src.cell(idx, 6).value = cap.source if cap else ""
            src.cell(idx, 7).value = cap.title if cap else ""
            src.cell(idx, 8).value = cap.observations if cap else ""
            for col in range(1, 9):
                c = src.cell(idx, col)
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = border
        src.freeze_panes = "A2"
        src.auto_filter.ref = f"A1:H{src.max_row}"
        for col, width in {"A": 9, "B": 70, "C": 24, "D": 42, "E": 14, "F": 22, "G": 60, "H": 70}.items():
            src.column_dimensions[col].width = width

    wb.save(output_path)
    return output_path


def write_log(captures: List[CaptureResult], config: Dict[str, Any]) -> Path:
    log_path = Path(config["output"]["log_csv"])
    ensure_dir(log_path.parent)
    with log_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["item", "status", "source", "brand", "image_path", "title", "observations", "url"])
        writer.writeheader()
        for cap in captures:
            writer.writerow(asdict(cap))
    return log_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def create_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"
    ws.append(["Item", "Descrição", "UN", "Qtde", "Valor Unitário", "Link", "Marca Manual (opcional)"])
    examples = [
        [1, "Conjunto de copos de vidro para água 350ml com 6un.", "Conjunto", 4, 42.05, "https://...", ""],
        [2, "Conjunto de taça de vinho em vidro com 6un", "Peça", 4, 87.40, "https://...", ""],
    ]
    for row in examples:
        ws.append(row)
    for col, width in {"A": 9, "B": 60, "C": 12, "D": 10, "E": 16, "F": 80, "G": 26}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    wb.save(path)


def main(argv: Optional[List[str]] = None) -> int:
    setup_logging()
    parser = argparse.ArgumentParser(description="Gera planilha com imagens reais de produtos a partir de links de fornecedores.")
    sub = parser.add_subparsers(dest="command", required=True)

    p_run = sub.add_parser("run", help="Executa a coleta e gera a planilha final.")
    p_run.add_argument("--config", default="config.example.yaml", help="Caminho do arquivo YAML de configuração.")

    p_template = sub.add_parser("template", help="Cria uma planilha modelo para futuras pesquisas.")
    p_template.add_argument("--output", default="modelo_pesquisa_produtos.xlsx", help="Arquivo Excel modelo de saída.")

    args = parser.parse_args(argv)

    if args.command == "template":
        create_template(Path(args.output))
        logging.info("Modelo criado em: %s", args.output)
        return 0

    if args.command == "run":
        config = load_config(Path(args.config))
        ensure_dir(Path(config["output"]["images_dir"]))
        ensure_dir(Path(config["output"]["workbook_path"]).parent)

        rows = read_memory_excel(config)
        rows = merge_pdf_data(rows, config)
        if not rows:
            logging.error("Nenhum item encontrado na planilha de entrada.")
            return 2

        captures = capture_all_images(rows, config)
        log_path = write_log(captures, config)
        output_path = build_output_workbook(rows, captures, config)

        ok = sum(1 for c in captures if c.status == "OK")
        revisar = len(captures) - ok
        logging.info("Planilha final: %s", output_path)
        logging.info("Log de execução: %s", log_path)
        logging.info("Itens com imagem OK: %s | Itens para revisar: %s", ok, revisar)
        return 0

    return 1


if __name__ == "__main__":
    raise SystemExit(main())
